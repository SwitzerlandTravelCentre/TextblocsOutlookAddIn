import argparse
import json
import re
import unicodedata
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles.colors import COLOR_INDEX


LANGUAGE_COLUMNS = [
    ("German", "DE", "German"),
    ("French", "FR", "French"),
    ("English", "EN", "English"),
    ("Englisch", "EN", "English"),
    ("Italian", "IT", "Italian"),
]

PARENTHESIZED_URL_PATTERN = re.compile(r"^\s*\((https?://[^)\s]+|mailto:[^)\s]+)\)", re.IGNORECASE)


def clean_header(value):
    return str(value or "").strip()


def clean_text(value):
    if value is None:
        return ""

    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    if len(text) >= 2 and text[0] == text[-1] == '"':
        text = text[1:-1].strip()

    return text


def parse_weight(value):
    if value is None or str(value).strip() == "":
        return 0

    try:
        number = float(str(value).strip().replace(",", "."))
    except ValueError:
        return 0

    return int(number) if number.is_integer() else number


def slug(value):
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", normalized).strip("-").lower()
    return normalized[:70] or "textblock"


def find_header_row(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=False):
        row_headers = [clean_header(cell.value) for cell in row]

        if "Category" in row_headers and "Topic" in row_headers:
            headers = {clean_header(cell.value): cell.column for cell in row if clean_header(cell.value)}
            return row[0].row, headers

    return None, {}


def color_to_rgb(color):
    if color is None:
        return None

    if color.type == "rgb" and color.rgb:
        return str(color.rgb)[-6:]

    if color.type == "indexed" and color.indexed is not None:
        try:
            return COLOR_INDEX[int(color.indexed)][-6:]
        except (IndexError, TypeError, ValueError):
            return None

    return None


def is_red_font(font):
    rgb = color_to_rgb(getattr(font, "color", None))

    if not rgb:
        return False

    try:
        red = int(rgb[0:2], 16)
        green = int(rgb[2:4], 16)
        blue = int(rgb[4:6], 16)
    except ValueError:
        return False

    return red >= 180 and green <= 110 and blue <= 110


def run_from_text(text, font=None, href=None):
    run = {"text": str(text or "")}

    if getattr(font, "b", False) or getattr(font, "bold", False):
        run["bold"] = True

    if is_red_font(font):
        run["highlight"] = "required"

    if getattr(font, "u", None) or getattr(font, "underline", None):
        run["underline"] = True

    if href:
        run["href"] = href

    return run


def cell_runs(cell):
    value = cell.value
    href = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else None

    if value is None:
        return []

    if isinstance(value, CellRichText):
        runs = []

        for item in value:
            if isinstance(item, TextBlock):
                runs.append(run_from_text(item.text, item.font))
            else:
                runs.append({"text": str(item)})

        return runs

    return [run_from_text(value, cell.font, href)]


def run_style(run):
    return tuple((key, run.get(key)) for key in ("bold", "highlight", "href", "underline") if run.get(key))


def merge_runs(runs):
    merged = []

    for run in runs:
        text = run.get("text", "")

        if not text:
            continue

        cleaned = {key: value for key, value in run.items() if value}
        cleaned["text"] = text

        if merged and run_style(merged[-1]) == run_style(cleaned):
            merged[-1]["text"] += text
        else:
            merged.append(cleaned)

    return merged


def normalize_runs(runs):
    characters = []

    for run in runs:
        text = run.get("text", "").replace("\r\n", "\n").replace("\r", "\n")
        style = {key: value for key, value in run.items() if key != "text" and value}

        for character in text:
            characters.append((character, style))

    start = 0
    end = len(characters)

    while start < end and characters[start][0].isspace():
        start += 1

    while end > start and characters[end - 1][0].isspace():
        end -= 1

    characters = characters[start:end]
    normalized = []

    for character, style in characters:
        if character == "\n":
            while normalized and normalized[-1][0] in " \t":
                normalized.pop()
            normalized.append((character, style))
            continue

        if character in " \t" and normalized and normalized[-1][0] == "\n":
            continue

        normalized.append((character, style))

    collapsed = []
    newline_count = 0

    for character, style in normalized:
        if character == "\n":
            newline_count += 1

            if newline_count <= 2:
                collapsed.append((character, style))
            continue

        newline_count = 0
        collapsed.append((character, style))

    output_runs = []

    for character, style in collapsed:
        run = {"text": character, **style}
        output_runs.append(run)

    return merge_runs(output_runs)


def plain_text_from_runs(runs):
    return "".join(run.get("text", "") for run in runs)


def sanitize_href(value):
    if not value:
        return None

    href = str(value).strip()

    if re.match(r"^(https?://|mailto:)", href, re.IGNORECASE):
        return href

    return None


def resolve_underlined_links(runs):
    resolved = []
    pending = deepcopy(runs)
    index = 0

    while index < len(pending):
        run = pending[index]

        if not run.get("underline"):
            resolved.append(run)
            index += 1
            continue

        label_runs = []

        while index < len(pending) and pending[index].get("underline"):
            label_run = dict(pending[index])
            label_run.pop("underline", None)
            label_runs.append(label_run)
            index += 1

        href = None

        if index < len(pending):
            match = PARENTHESIZED_URL_PATTERN.match(pending[index].get("text", ""))

            if match:
                href = sanitize_href(match.group(1))
                pending[index]["text"] = pending[index]["text"][match.end() :]

                if not pending[index]["text"]:
                    index += 1

        if href:
            for label_run in label_runs:
                label_run["href"] = href
                resolved.append(label_run)
        else:
            resolved.extend(label_runs)

    return merge_runs(resolved)


def formatted_runs_for_json(runs):
    json_runs = []

    for run in runs:
        text = run.get("text", "")

        if not text:
            continue

        json_run = {"text": text}

        if run.get("bold"):
            json_run["bold"] = True

        if run.get("highlight") == "required":
            json_run["highlight"] = "required"

        href = sanitize_href(run.get("href"))

        if href:
            json_run["href"] = href

        json_runs.append(json_run)

    return merge_runs(json_runs)


def cell_text_and_formatting(cell):
    normalized_runs = normalize_runs(cell_runs(cell))
    text = plain_text_from_runs(normalized_runs)

    if not text:
        return "", None

    linked_runs = resolve_underlined_links(normalized_runs)
    formatted_runs = formatted_runs_for_json(linked_runs)
    has_formatting = any(any(key in run for key in ("bold", "highlight", "href")) for run in formatted_runs)

    return text, formatted_runs if has_formatting else None


def convert(source_path, output_path):
    workbook = openpyxl.load_workbook(source_path, data_only=True, rich_text=True)
    blocks = []
    source_rows = 0

    for sheet in workbook.worksheets:
        header_row_index, headers = find_header_row(sheet)

        if header_row_index is None:
            continue

        category_column = headers.get("Category")
        topic_column = headers.get("Topic")
        usage_column = headers.get("Usage")
        weight_column = headers.get("Weight") or headers.get("weight")
        language_map = []
        seen_language_codes = set()

        for field, code, label in LANGUAGE_COLUMNS:
            if field in headers and code not in seen_language_codes:
                language_map.append((field, code, label, headers[field]))
                seen_language_codes.add(code)

        for row_index in range(header_row_index + 1, sheet.max_row + 1):
            category = clean_text(sheet.cell(row_index, category_column).value) if category_column else ""
            topic = clean_text(sheet.cell(row_index, topic_column).value) if topic_column else ""
            usage = clean_text(sheet.cell(row_index, usage_column).value) if usage_column else ""
            weight = parse_weight(sheet.cell(row_index, weight_column).value) if weight_column else 0
            language_values = []

            for field, code, label, column in language_map:
                text, formatted_runs = cell_text_and_formatting(sheet.cell(row_index, column))
                language_values.append((field, code, label, text, formatted_runs))

            if not category and not topic and not usage and not any(text for *_, text, _formatted in language_values):
                continue

            source_rows += 1
            base_id = f"{slug(category)}-{slug(topic)}-{row_index}"

            for _field, code, label, text, formatted_runs in language_values:
                if not text:
                    continue

                block = {
                    "id": f"{base_id}-{code.lower()}",
                    "sourceItemId": f"{sheet.title}!{row_index}",
                    "sourceSheet": sheet.title,
                    "sourceRow": row_index,
                    "category": category,
                    "topic": topic,
                    "usage": usage,
                    "language": code,
                    "languageLabel": label,
                    "weight": weight,
                    "text": text,
                }

                if formatted_runs:
                    block["formattedText"] = formatted_runs

                blocks.append(block)

    payload = {
        "schemaVersion": 1,
        "source": {
            "type": "xlsx",
            "fileName": source_path.name,
            "sheetCount": len(workbook.worksheets),
            "dataRowCount": source_rows,
        },
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "blockCount": len(blocks),
        "blocks": blocks,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    return payload


def main():
    parser = argparse.ArgumentParser(description="Convert the STC email text block workbook to JSON.")
    parser.add_argument("source", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("public/data/textblocks.json"),
    )
    args = parser.parse_args()
    payload = convert(args.source, args.output)
    print(json.dumps({"rows": payload["source"]["dataRowCount"], "blocks": payload["blockCount"]}))


if __name__ == "__main__":
    main()
